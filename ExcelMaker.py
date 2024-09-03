import tkinter as tk
import xlsxwriter as xls
import calendar 
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment,Border, Side, Font, PatternFill, Alignment

year = 2024 #Select the year 
month = 9 #Select the month
name = 'test'
cl = calendar.Calendar()

#This function will help in order to put the montly summary correct
def swich(value):
    if value == 1:
        return str('G')
    elif value == 2:
        return str('H')
    elif value == 3:
        return str('I')
    elif value == 4:
        return str('J')
    elif value == 5:
        return str('K')
    
#This fucntion helps to create a sheet
def sheet(day):
    
    ws2 = wb.create_sheet(title=str(day))
   
    
    # Registro de Dia 
    ws2['E1'] = 'REGISTRO DE DIA'
    ws2.merge_cells('E1:F2')
    ws2['E1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['E1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    theborder = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))

    fontTitles = Font(name='Arial', color="F09B30", bold=True)
    fontSums = Font(name='Arial', color="000000", bold=True)
    ws2['E1'].font = fontTitles
    
    def titulos(celda, descripcion, size, money):
        ws2[celda] = descripcion
        ws2[celda].alignment = Alignment(horizontal='center', vertical='center')
        ws2[celda].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws2[celda].font = fontTitles
        letter = celda[0:1]
        ws2.column_dimensions[letter].width = size
        ws2[celda].border = theborder
        
        for number in range(4, 41):
            cell_coordinate = f"{letter}{number}"
            if money:
                ws2[cell_coordinate].number_format = '$#,##0.00'
                
                ws2[cell_coordinate].font = fontSums
            ws2[cell_coordinate].border = theborder
        
    # Titulos de tablas
    titulos('B3', 'ID', 8, False)
    titulos('C3', 'CONCEPTO', 12, False)
    titulos('D3', 'MEMBRESIA EFECTIVO', 35, True)
    titulos('E3', 'MEMBRESIA TARJETA', 35, True)
    titulos('F3', 'PRODUCTO EFECTIVO', 35, True)
    titulos('G3', 'PRODUCTO TARJETA', 35, True)
    titulos('H3', 'FONDEADORA', 35, True)
    titulos('I3', 'COMENTARIOS', 20, False)
    titulos('C41', 'TOTAL', 12, False)
    
    for number in range(4, 41):
        cell_coordinate = f"{'B'}{number}" 
        ws2[cell_coordinate].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
        ws2[cell_coordinate].font = Font(name='Arial', color="000000", bold=False)
         
    def sumas(celda):
        letter = celda[0:1]
        ws2[celda] = '=SUM(' + letter + str(4) + ':' + letter + str(40) + ')'
        ws2[celda].data_type = 'f'
        ws2[celda].alignment = Alignment(horizontal='center', vertical='center')
        ws2[celda].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
        ws2[celda].number_format = '$#,##0.00'
        ws2[celda].border = theborder
        
    sumas('D41')
    sumas('E41')
    sumas('F41')
    sumas('G41')
    sumas('H41')
    
    def sensoDebito(letra,celda):
        
        
        ws2[celda].alignment = Alignment(horizontal='center', vertical='center')
        ws2[celda].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
        ws2[celda].number_format = '$#,##0.00'
        ws2[celda].border = theborder
        ws2[celda] = f'=SUMAPRODUCTO((IZQUIERDA(C4:C40)=\"{letra}\")*D4:D40)'#This was part of a new fucntion that was implemented in the second version, However it has a few mistakes with excel in which you have to replace a @ with "" in all the pages of the Excel Book
        ws2[celda].data_type = 'f'
    
    def sensoTarjeta(letra,celda):
      
        ws2[celda].alignment = Alignment(horizontal='center', vertical='center')
        ws2[celda].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
        ws2[celda].number_format = '$#,##0.00'
        ws2[celda].border = theborder
        ws2[celda].data_type = 'f'
        ws2[celda] =  f'=SUMAPRODUCTO((IZQUIERDA(C4:C40)=\"{letra}\")*E4:E40)' #This was part of a new fucntion that was implemented in the second version, However it has a few mistakes with excel in which you have to replace a @ with "" in all the pages of the Excel Book
        
    
   
    
   
    titulos('H51', 'ZUMBA EFECTIVO', 35, False)
    sensoDebito('Z', 'I51')
    titulos('H52', 'ZUMBA TARJETA', 35, False)
    sensoTarjeta('Z', 'I52')
    titulos('H53', 'Crossfit EFECTIVO', 35, False)
    sensoDebito('C', 'I53')
    titulos('H54', 'Crossfit TARJETA ', 35, False)
    sensoTarjeta('C', 'I54')
    
    
    titulos('H45', 'MEMBRESIA EFECTIVO', 35, False)
    titulos('H46', 'MEMBRESIA TARJETA', 35, False)
    titulos('H47', 'PRODUCTO EFECTIVO', 35, False)
    titulos('H48', 'PRODUCTO TARJETA', 35, False)
    titulos('H49', 'FONDEADORA', 35, False)
    titulos('H50', 'TOTAL: ', 35, False)
    
    def final(consult, result):
        ws2[result] = '=' + consult
        ws2[result].data_type = 'f'
        ws2[result].alignment = Alignment(horizontal='center', vertical='center')
        ws2[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
        ws2[result].number_format = '$#,##0.00'
        ws2[result].border = theborder
        
    final('D41', 'I45')
    final('E41', 'I46') 
    final('F41', 'I47') 
    final('G41', 'I48') 
    final('H41', 'I49') 

    ws2['I50'] = '=SUM(I45:I49)'
    ws2['I50'].data_type = 'f'
    ws2['I50'].alignment = Alignment(horizontal='center', vertical='center')
    ws2['I50'].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
    ws2['I50'].number_format = '$#,##0.00'
    ws2['I50'].border = theborder
    
     
   
    return 0
#This will create the number of sheets necesary in order to have the  correct sheet  amount
def sheetMaker(year, month):
    myList = [cl.monthdays2calendar(year, month)]
    for i, sublist in enumerate(myList):
        count = len(sublist)
    day = 1
    for i in range(1):
        for j in range(count):  # Range from 0 to 5 
            
            list =  []
            for k  in range(1,7):  # Days Monday tru Saturday 
          
                    
                my_tuple =myList[i][j][k-1]
                
                
                if (my_tuple[0] !=0) : #Sundays are not part of the monthly days so it cannot be 0 
                        sheet(day)
                        list.append(day)
                        day += 1
                                    
            day+=1 
            
            formula = '=' 
            dataF1 =formula
            dataF2 =formula
            dataF3 =formula
            dataF4 =formula
            dataF5 =formula
            dataFZE = formula
            dataFZC = formula
            dataFCE = formula  
            dataFCC = formula
            for x in range(len(list)):
                
                data1 = f"'{list[x]}'!I45 "
                data2 = f"'{list[x]}'!I46 "
                data3 = f"'{list[x]}'!I47 "
                data4 = f"'{list[x]}'!I48 "
                data5 = f"'{list[x]}'!I49 "
                
                dataZE = f"'{list[x]}'!I51"
                dataZC = f"'{list[x]}'!I52"
                dataCE = f"'{list[x]}'!I53"
                dataCC = f"'{list[x]}'!I54"
                
                dataF1 = dataF1 + data1 
                dataF2 = dataF2 + data2
                dataF3 = dataF3 + data3
                dataF4 = dataF4 + data4
                dataF5 = dataF5 + data5
                #DATA FORM THE CLASSES
                dataFZE = dataFZE + dataZE
                dataFZC = dataFZC + dataZC 
                dataFCE = dataFCE + dataCE  
                dataFCC = dataFCC + dataCC
                 
                if list[x] != list[len(list)-1]:
                    dataF1 += '+'
                    dataF2 += '+'
                    dataF3 += '+'
                    dataF4 += '+'
                    dataF5 += '+'
                    dataFZE += '+'
                    dataFZC += '+'
                    dataFCE += '+'
                    dataFCC += '+'
                    
                   
           
            #this is the resume page, in which you can see everything related to the clases, a quick summary of the month
            word = str(swich(j+1))
            value = '10'
            result = word + value
            
            
            ws[result] = dataF1   
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
            
            value = '11'
            result =word+value
            ws[result] =dataF2  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
            
            value = '12'
            result =word+value
            ws[result] =dataF3  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
            
            value = '13'
            result =word+value      
            ws[result] =dataF4  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
            
            value = '14'
            result =word+value 
            ws[result] =dataF5  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder 
            
            value = '20'
            result =word+value 
            ws[result] =dataFZE  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder 
            
            value = '21'
            result =word+value 
            ws[result] =dataFZC  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
            
            value = '22'
            result =word+value 
            ws[result] =dataFCE  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
            
            value = '23'
            result =word+value 
            ws[result] =dataFCC  
            ws[result].data_type = 'f'
            ws[result].alignment = Alignment(horizontal='center', vertical='center')
            ws[result].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
            ws[result].number_format = '$#,##0.00'
            ws[result].border = theborder
        #MAKE A FUNCTION HERE TO REDUCE THE AMOUNT OF LINES
            
            

   
      
   

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Resumen Mensual"  

ws['G3'] = 'RESUMEN MENSUAL'
ws.merge_cells('G3:L7')
ws['G3'].alignment = Alignment(horizontal='center', vertical='center')
ws['G3'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
theborder = Border(left=Side(style='medium'),
                       right=Side(style='medium'),
                       top=Side(style='medium'),
                       bottom=Side(style='medium'))

fontTitles = Font(name='Arial', color="F09B30", bold=True)
fontSums = Font(name='Arial', color="000000", bold=True)
ws['G3'].font = fontTitles

def  titulos(descripcion,celda):
    ws[celda] = descripcion
    ws[celda].alignment = Alignment(horizontal='center', vertical='center')
    ws[celda].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws[celda].font = fontTitles
    letter = celda[0:1]
    ws.column_dimensions[letter].width = 15
    ws[celda].border = theborder
    
def subtitulos(descripcion, celda):
    ws[celda] = descripcion
    ws[celda].alignment = Alignment(horizontal='center', vertical='center')
    ws[celda].fill = PatternFill(start_color="F09B30", end_color="F09B30", fill_type="solid")
    ws[celda].font = fontSums
    ws.column_dimensions['F'].width = 30
    ws[celda].border = theborder

titulos('Semana 1','G9')
titulos('Semana 2','H9')
titulos('Semana 3','I9')
titulos('Semana 4','J9')
titulos('Semana 5','K9')
titulos('Total', 'L9')
titulos(' ','F9')
subtitulos('Total de Membresias Efectivo: ', 'F10')
subtitulos('Total de Membresias Tarjeta: ', 'F11')
subtitulos('Total de Producto Efectivo: ', 'F12')
titulos('Total de semana: ','F15')
subtitulos('Total de Producto Tarjeta: ', 'F13')
subtitulos('Transferencia ', 'F14')
titulos('Semana 1','G19')
titulos('Semana 2','H19')
titulos('Semana 3','I19')
titulos('Semana 4','J19')
titulos('Semana 5','K19')
titulos('Total', 'L19')
titulos('Ingreso por clase','F19')
subtitulos('Total Zumba en efectivo: ', 'F20')
subtitulos('Total de Zumba en tarjeta: ', 'F21')
subtitulos('Total de Crossfit efectivo: ', 'F22')
subtitulos('Total de Crossfit Tarjeta : ', 'F23')
titulos('Total de semana: ','F24')

#MAKE A FUNCTION DOWN HERE


def titleSum(cell,sum,orange):
    
    color ="000000"
    font = fontTitles
    if(orange):
        color = "F09B30"
        font = fontSums
    
    ws[cell] = sum
    ws[cell].data_type = 'f'
    ws.column_dimensions['f'].width = 30
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws[cell].font = font
    ws[cell].number_format = '$#,##0.00'
    ws[cell].border = theborder
    
    



titleSum('G24','=SUM(G20:G23)',False)
titleSum('H24','=SUM(H20:H23)',False)
titleSum('I24','=SUM(I20:I23)',False)
titleSum('J24','=SUM(J20:J23)',False)
titleSum('K24','=SUM(K20:K23)',False)


titleSum('L20','=SUM(G20:K20)',True)
titleSum('L21','=SUM(G21:K21)',True)
titleSum('L22','=SUM(G22:K22)',True)
titleSum('L23','=SUM(G23:K23)',True)
titleSum('L24','=SUM(G24:K24)',False)

titleSum('G15','=SUM(G10:G14)',False)
titleSum('H15','=SUM(H10:H14)',False)
titleSum('I15','=SUM(I10:I14)',False)
titleSum('J15','=SUM(J10:J14)',False)
titleSum('K15','=SUM(K10:K14)',False)
titleSum('L15','=SUM(L10:L14)',False)


titleSum('L10','=SUM(G10:K10)',True)
titleSum('L11','=SUM(G11:K11)',True)
titleSum('L12','=SUM(G12:K12)',True)
titleSum('L13','=SUM(G13:K13)',True)
titleSum('L14','=SUM(G14:K14)',True)


sheetMaker(year,month)


directory = '' #THIS WILL HELP TO SELECT WHERE YOU WANT TO PUT THE EXCEL
file_name = 'Septiembre2024.xlsx' #NAME OF THE FILE
file_path = os.path.join(directory, file_name)

    # Ensure the directory exists
os.makedirs(directory, exist_ok=True)

    # Save the workbook to the specified directory
wb.save(file_path) #SAVE 


